"""Apply a vendor wholesale price list to catalog_items.

For each catalog row of the selected vendor that matches a price-list
entry on ``style_number``, records the wholesale cost and recomputes the
full-package shelf price via ``services.pricing`` into
``unit_price_cents``, plus provenance (as-of date and a source label).

Each vendor publishes prices in a different workbook shape, so the only
vendor-specific piece is the reader (see ``VENDOR_PROFILES``); the
compute / diff / DB-apply core is shared:

  - ``morilee``: consolidated workbook, one tab per line; the Quince tab
    is read by default; columns Style Number / .. / WS / MSRP; the as-of
    date is parsed from the tab name. This is an all-styles list.
  - ``rachel_allan``: a single-SEASON line sheet (e.g. Fall 2026). All
    sheets are scanned; a data row has a style in col 0 and wholesale in
    the SALE PRICE column (col 6); no MSRP, no as-of. NOTE: a season line
    sheet only contains that season's NEW styles, so it covers a subset
    of an actively-sold catalog — carryover styles need their own
    season's sheet (re-run; the apply is idempotent and stacks).

Two modes:

  - **default (dry-run)**: compute everything, write nothing, emit a
    review report (old price -> new price, wholesale, multiplier, as-of).
  - **--apply**: commit the wholesale + shelf-price updates.

Pricing is per style, so every color row of a style gets the same
wholesale and shelf price. Rows whose wholesale is below the $299 floor
are reported as ``needs_manual_pricing`` and left untouched. Historical
quotes/invoices snapshot their own price on their line rows, so applying a
new list never changes past documents.

Usage:

    venv/bin/python scripts/seed_catalog/apply_price_list.py --vendor morilee
    venv/bin/python scripts/seed_catalog/apply_price_list.py \\
        --vendor rachel_allan --apply
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(_REPO_ROOT / ".env")
os.environ.setdefault("APP_TIMEZONE", "America/Chicago")

import openpyxl  # noqa: E402

from database.connection import SessionLocal  # noqa: E402
from database.models import CatalogItem  # noqa: E402
from services.pricing import (  # noqa: E402
    base_multiplier,
    calculate_dress_price,
    parse_tab_as_of,
)

DEFAULT_SUMMARY_DIR = _REPO_ROOT / "data/reports"


def _norm_style(value: Any) -> str:
    """Normalize a style cell to the catalog's string form.

    Style numbers are stored as floats in the workbook (e.g. 4080040.0);
    strip the trailing ``.0`` so they match ``catalog_items.style_number``.
    """
    if value is None:
        return ""
    s = str(value).strip()
    return s[:-2] if s.endswith(".0") else s


def _to_cents(value: Any) -> int | None:
    """Dollars (int/float or a "$1,279" string) -> integer cents."""
    if value is None or value == "":
        return None
    s = str(value).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def _cents(v: int | None) -> str:
    return "—" if v is None else f"${v / 100:,.2f}"


# ---------------------------------------------------------------------------
# Vendor readers
#
# Each vendor publishes its wholesale prices in a different workbook shape,
# so the *reader* is the only vendor-specific piece. Every reader returns a
# normalized PriceList; the compute / diff / DB-apply core below is shared.
# ---------------------------------------------------------------------------


@dataclass
class PriceList:
    # style_number -> {wholesale_cents, msrp_cents, name}
    prices: dict[str, dict[str, Any]]
    as_of: date | None
    source_label: str
    designer: str
    label: str  # which tab(s) were read, for the report header


def _select_tab(wb: "openpyxl.Workbook", tab: str | None, prefix: str) -> str:
    """Pick a worksheet by explicit name, else the first tab starting with
    ``prefix`` that isn't an accessories tab."""
    if tab:
        if tab not in wb.sheetnames:
            raise SystemExit(f"tab {tab!r} not found; tabs: {wb.sheetnames}")
        return tab
    for name in wb.sheetnames:
        low = name.lower()
        if low.startswith(prefix.lower()) and "accessor" not in low:
            return name
    raise SystemExit(
        f"no tab starting with {prefix!r} found; pass --tab. "
        f"tabs: {wb.sheetnames}"
    )


def _read_morilee(path: Path, tab: str | None) -> PriceList:
    """Morilee consolidated list: one tab per line, columns Style Number,
    Style Name, Season, Colors, WS, MSRP. As-of date is in the tab name."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    actual_tab = _select_tab(wb, tab, "Quince as of")
    rows = list(wb[actual_tab].iter_rows(values_only=True))
    header_idx = next(
        (
            i
            for i, r in enumerate(rows)
            if r and str(r[0] or "").strip().lower() == "style number"
        ),
        None,
    )
    if header_idx is None:
        raise SystemExit(f"no 'Style Number' header found on tab {actual_tab!r}")
    prices: dict[str, dict[str, Any]] = {}
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None:
            continue
        style = _norm_style(r[0])
        if not style:
            continue
        prices[style] = {
            "name": r[1] if len(r) > 1 else None,
            "wholesale_cents": _to_cents(r[4] if len(r) > 4 else None),
            "msrp_cents": _to_cents(r[5] if len(r) > 5 else None),
        }
    return PriceList(
        prices=prices,
        as_of=parse_tab_as_of(actual_tab),
        source_label=f"Morilee Consolidated / {actual_tab}",
        designer="Morilee",
        label=actual_tab,
    )


# Rachel Allan style codes: optional 1-3 letter prefix + 3-5 digits +
# optional trailing letter (RB1020, 40543, RQ1190, OVL1002, 40656C).
_RA_STYLE_RE = re.compile(r"^[A-Z]{0,3}\d{3,5}[A-Z]?$")


# A Rachel Allan PDF price-list row: style code, then the wholesale price
# and (optionally) the MSRP as the trailing dollar amounts on the line.
_RA_PDF_ROW = re.compile(
    r"(?m)^\s*([A-Z]{0,3}\d{3,6}[A-Z]?)\b.*?\$([\d,]+)(?:\s+\$([\d,]+))?\s*$"
)


def _pdf_text(path: Path) -> str:
    """Extract text from a PDF, preferring poppler's pdftotext (keeps the
    column layout), falling back to pypdf."""
    try:
        out = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            capture_output=True,
            text=True,
        )
        if out.returncode == 0 and out.stdout.strip():
            return out.stdout
    except FileNotFoundError:
        pass
    import pypdf

    reader = pypdf.PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _read_rachel_allan_pdf(path: Path) -> PriceList:
    """Rachel Allan PDF price list: each row is `STYLE ... $WHOLESALE
    $MSRP`. The first dollar amount is wholesale (the MSRP, when present,
    is ~2x and used only as a reference)."""
    prices: dict[str, dict[str, Any]] = {}
    for m in _RA_PDF_ROW.finditer(_pdf_text(path)):
        ws_cents = _to_cents(m.group(2))
        if ws_cents is None:
            continue
        prices[m.group(1)] = {
            "name": None,
            "wholesale_cents": ws_cents,
            "msrp_cents": _to_cents(m.group(3)) if m.group(3) else None,
        }
    return PriceList(
        prices=prices,
        as_of=None,
        source_label=f"Rachel Allan / {path.name}",
        designer="Rachel Allan",
        label=path.name,
    )


def _read_rachel_allan(path: Path, tab: str | None) -> PriceList:
    """Rachel Allan price list. PDFs (full/master lists) and the .xlsx
    season line sheet are both supported.

    xlsx line sheet: every sheet is scanned, skipping the repeating section
    banners and header rows; a data row has a style code in col 0 and the
    wholesale figure in the SALE PRICE column (col 6). No MSRP, no as-of.
    A season line sheet only carries that season's NEW styles."""
    if path.suffix.lower() == ".pdf":
        return _read_rachel_allan_pdf(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [tab] if tab else list(wb.sheetnames)
    prices: dict[str, dict[str, Any]] = {}
    for name in sheets:
        for r in wb[name].iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            style = str(r[0]).strip()
            if not _RA_STYLE_RE.match(style):
                continue
            cents = _to_cents(r[6] if len(r) > 6 else None)
            if cents is None:
                continue
            prices[style] = {
                "name": None,
                "wholesale_cents": cents,
                "msrp_cents": None,
            }
    return PriceList(
        prices=prices,
        as_of=None,
        source_label="Rachel Allan / Fall 2026 line sheet",
        designer="Rachel Allan",
        label=" + ".join(sheets),
    )


VENDOR_PROFILES: dict[str, dict[str, Any]] = {
    "morilee": {
        "reader": _read_morilee,
        "default_xlsx": _REPO_ROOT
        / "Morilee New York Consolidated Price List _ As Of November 2025.xlsx",
    },
    "rachel_allan": {
        "reader": _read_rachel_allan,
        "default_xlsx": _REPO_ROOT / "RA Fall line Sheet 2026.xlsx",
    },
}


def apply_price_list(
    *,
    vendor: str,
    xlsx_path: Path,
    tab: str | None,
    apply: bool,
) -> dict[str, Any]:
    reader = VENDOR_PROFILES[vendor]["reader"]
    pl = reader(xlsx_path, tab)
    price_list = pl.prices
    as_of = pl.as_of
    source_label = pl.source_label
    designer = pl.designer
    actual_tab = pl.label

    updated: list[dict[str, Any]] = []
    unchanged: list[str] = []
    needs_manual: list[dict[str, Any]] = []
    no_price_row: list[str] = []  # catalog rows with no matching style in list

    db = SessionLocal()
    try:
        rows = (
            db.query(CatalogItem)
            .filter(CatalogItem.designer == designer)
            .all()
        )
        for item in rows:
            style = (item.style_number or "").strip()
            entry = price_list.get(style)
            if entry is None:
                no_price_row.append(item.internal_sku)
                continue

            ws_cents = entry["wholesale_cents"]
            if ws_cents is None:
                no_price_row.append(item.internal_sku)
                continue

            result = calculate_dress_price(ws_cents)
            mult = base_multiplier(ws_cents)

            if result.requires_manual_pricing:
                needs_manual.append(
                    {
                        "internal_sku": item.internal_sku,
                        "style_number": style,
                        "color": item.color,
                        "wholesale_cents": ws_cents,
                        "reason": "below $299 pricing floor",
                    }
                )
                continue

            new_price = result.final_price_cents
            old_price = item.unit_price_cents
            old_ws = item.wholesale_cents
            price_changed = old_price != new_price
            ws_changed = old_ws != ws_cents
            asof_changed = item.wholesale_as_of != as_of

            if not (price_changed or ws_changed or asof_changed):
                unchanged.append(item.internal_sku)
                continue

            change = {
                "internal_sku": item.internal_sku,
                "style_number": style,
                "color": item.color,
                "wholesale_old_cents": old_ws,
                "wholesale_new_cents": ws_cents,
                "multiplier": mult,
                "unit_price_old_cents": old_price,
                "unit_price_new_cents": new_price,
                "msrp_reference_cents": entry["msrp_cents"],
                "as_of": as_of.isoformat() if as_of else None,
            }
            updated.append(change)

            if apply:
                item.wholesale_cents = ws_cents
                item.wholesale_as_of = as_of
                item.wholesale_source = source_label
                item.unit_price_cents = new_price
                item.updated_at = datetime.now(timezone.utc)

        if apply:
            db.commit()
        else:
            db.rollback()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "vendor": vendor,
        "xlsx": str(xlsx_path.name),
        "tab": actual_tab,
        "as_of": as_of.isoformat() if as_of else None,
        "source_label": source_label,
        "designer": designer,
        "mode": "apply" if apply else "dry-run",
        "price_list_styles": len(price_list),
        "counts": {
            "updated": len(updated),
            "unchanged": len(unchanged),
            "needs_manual_pricing": len(needs_manual),
            "catalog_rows_without_price": len(no_price_row),
        },
        "updated": updated,
        "needs_manual_pricing": needs_manual,
        "catalog_rows_without_price": no_price_row,
    }


def print_summary(summary: dict[str, Any]) -> None:
    print(
        f"price list apply [{summary['vendor']}]: {summary['xlsx']}  "
        f"tab='{summary['tab']}'  as-of={summary['as_of']}  "
        f"mode={summary['mode']}"
    )
    c = summary["counts"]
    print(f"  price-list styles:           {summary['price_list_styles']}")
    print(f"  catalog rows updated:        {c['updated']}")
    print(f"  unchanged:                   {c['unchanged']}")
    print(f"  needs manual pricing:        {c['needs_manual_pricing']}")
    print(f"  catalog rows w/o price row:  {c['catalog_rows_without_price']}")

    if summary["updated"]:
        print("\n  style    color                 wholesale     x mult   "
              "old price  -> new price     (MSRP ref)")
        for ch in summary["updated"][:40]:
            print(
                f"  {ch['style_number']:<8} {str(ch['color'])[:20]:<20} "
                f"{_cents(ch['wholesale_new_cents']):>10}  x{ch['multiplier']:<5} "
                f"{_cents(ch['unit_price_old_cents']):>10} -> "
                f"{_cents(ch['unit_price_new_cents']):>11}  "
                f"({_cents(ch['msrp_reference_cents'])})"
            )
        if len(summary["updated"]) > 40:
            print(f"  ... and {len(summary['updated']) - 40} more (see JSON)")

    if summary["needs_manual_pricing"]:
        print("\n  NEEDS MANUAL PRICING (below $299 floor):")
        for m in summary["needs_manual_pricing"]:
            print(
                f"    {m['style_number']} {m['color']}  "
                f"wholesale {_cents(m['wholesale_cents'])}"
            )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--vendor",
        choices=sorted(VENDOR_PROFILES),
        default="morilee",
        help="Which vendor price list to apply. Selects the reader and the "
        "default workbook + catalog designer.",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Workbook to read. Defaults to the selected vendor's file.",
    )
    parser.add_argument(
        "--tab",
        default=None,
        help="Worksheet tab to read. Morilee: defaults to the 'Quince as "
        "of ...' tab. Rachel Allan: defaults to ALL sheets.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Commit the updates. Without this flag the run is a dry-run.",
    )
    parser.add_argument(
        "--summary-output",
        type=Path,
        default=None,
        help="Report JSON path. Defaults to "
        "data/reports/price_list_apply_<vendor>.json.",
    )
    args = parser.parse_args()

    xlsx = args.xlsx or VENDOR_PROFILES[args.vendor]["default_xlsx"]
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")
    summary_output = args.summary_output or (
        DEFAULT_SUMMARY_DIR / f"price_list_apply_{args.vendor}.json"
    )

    summary = apply_price_list(
        vendor=args.vendor,
        xlsx_path=xlsx,
        tab=args.tab,
        apply=args.apply,
    )

    summary_output.parent.mkdir(parents=True, exist_ok=True)
    summary_output.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print_summary(summary)
    print(f"\nsummary written: {summary_output}")
    if not args.apply:
        print("DRY RUN — no changes written. Re-run with --apply to commit.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
